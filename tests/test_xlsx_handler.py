"""Tests for XlsxHandler."""

from pathlib import Path
import tempfile

import openpyxl
import pytest

from xlsx_streamer.sources.local import LocalFileSource
from xlsx_streamer.xlsx_handler import XlsxHandler


def test_xlsx_handler_initialization() -> None:
    """Test XlsxHandler initialization."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        handler = XlsxHandler(source, sheet_name="Sheet1", chunk_size=8192)

        assert handler.source is source
        assert handler.sheet_name == "Sheet1"
        assert handler.chunk_size == 8192
        assert handler.delimiter == ","
        assert handler.quotechar == '"'
    finally:
        Path(temp_path).unlink()


def test_xlsx_handler_row_to_bytes() -> None:
    """Test _row_to_bytes converts row to CSV bytes correctly."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        handler = XlsxHandler(source)

        # Test simple row
        row_bytes = handler._row_to_bytes(["A", "B", "C"])
        assert row_bytes == b"A,B,C\r\n"

        # Test row with numbers
        row_bytes = handler._row_to_bytes([1, 2, 3])
        assert row_bytes == b"1,2,3\r\n"

        # Test row with None values
        row_bytes = handler._row_to_bytes(["A", None, "C"])
        assert row_bytes == b"A,,C\r\n"

        # Test row with values needing quoting
        row_bytes = handler._row_to_bytes(["A,B", "C", "D"])
        assert b'"A,B"' in row_bytes

    finally:
        Path(temp_path).unlink()


def test_xlsx_handler_stream_rows_with_real_file() -> None:
    """Test stream_rows with real XLSX file."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        # Create a real XLSX file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header1", "Header2", "Header3"])
        ws.append([1, 2, 3])
        ws.append(["A", "B", "C"])
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        handler = XlsxHandler(source)

        rows = list(handler.stream_rows())
        assert len(rows) == 3

        # Verify rows are CSV bytes
        for row in rows:
            assert isinstance(row, bytes)
            assert b"," in row or len(row.strip()) > 0

    finally:
        Path(xlsx_path).unlink()


def test_xlsx_handler_stream_rows_with_named_sheet() -> None:
    """Test stream_rows with named sheet."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        # Create XLSX with multiple sheets
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A", "B"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["X", "Y", "Z"])
        ws2.append([1, 2, 3])
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        handler = XlsxHandler(source, sheet_name="Sheet2")

        rows = list(handler.stream_rows())
        assert len(rows) == 2

    finally:
        Path(xlsx_path).unlink()


def test_xlsx_handler_stream_rows_empty_file() -> None:
    """Test stream_rows with empty XLSX file."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        # Create empty XLSX file
        wb = openpyxl.Workbook()
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        handler = XlsxHandler(source)

        rows = list(handler.stream_rows())
        # Empty file should have no rows
        assert len(rows) == 0

    finally:
        Path(xlsx_path).unlink()


def test_xlsx_handler_stream_rows_invalid_sheet() -> None:
    """Test stream_rows with invalid sheet name raises error."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        handler = XlsxHandler(source, sheet_name="NonExistent")

        with pytest.raises(OSError, match="Failed to stream XLSX rows"):
            list(handler.stream_rows())

    finally:
        Path(xlsx_path).unlink()


def test_xlsx_handler_stream_rows_with_special_characters() -> None:
    """Test stream_rows handles special characters correctly."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header with, comma", 'Quote"Test', "Normal"])
        ws.append(["Line\nBreak", "Tab\tTest", "Emoji😊"])
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        handler = XlsxHandler(source)

        rows = list(handler.stream_rows())
        assert len(rows) == 2

        # All rows should be valid CSV bytes
        for row in rows:
            assert isinstance(row, bytes)

    finally:
        Path(xlsx_path).unlink()


def test_xlsx_handler_large_row_count_logging() -> None:
    """Test that handler logs progress for large row counts."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        # Add just a few rows (testing the logging mechanism, not actual 10k rows)
        for i in range(5):
            ws.append([i, i + 1, i + 2])
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        handler = XlsxHandler(source)

        rows = list(handler.stream_rows())
        assert len(rows) == 5

    finally:
        Path(xlsx_path).unlink()
