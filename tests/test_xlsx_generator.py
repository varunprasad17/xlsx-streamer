"""Tests for StreamingXlsxReader."""

from collections.abc import Iterator
import contextlib
from pathlib import Path
import tempfile

import openpyxl
import pytest

from xlsx_streamer.sources.local import LocalFileSource
from xlsx_streamer.xlsx_generator import IterableToFile, StreamingXlsxReader


def test_streaming_xlsx_reader_initialization() -> None:
    """Test StreamingXlsxReader initialization."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        reader = StreamingXlsxReader(
            source, target_worksheet_filepath="xl/worksheets/sheet1.xml", chunk_size=8192
        )

        assert reader.source is source
        assert reader.chunk_size == 8192
        assert reader.worksheet_filepath == "xl/worksheets/sheet1.xml"
    finally:
        Path(temp_path).unlink()


def test_streaming_xlsx_reader_default_worksheet() -> None:
    """Test StreamingXlsxReader uses default worksheet path."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        reader = StreamingXlsxReader(source)

        assert reader.worksheet_filepath == "xl/worksheets/sheet1.xml"
    finally:
        Path(temp_path).unlink()


def test_streaming_xlsx_reader_stream_rows() -> None:
    """Test stream_rows with real XLSX file."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B", "C"])
        ws.append([1, 2, 3])
        ws.append(["X", "Y", "Z"])
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        reader = StreamingXlsxReader(source, target_worksheet_filepath="xl/worksheets/sheet1.xml")

        rows = list(reader.stream_rows([]))
        assert len(rows) == 3

        # Check row content
        assert rows[0] == ["A", "B", "C"]
        assert rows[1] == [1, 2, 3]
        assert rows[2] == ["X", "Y", "Z"]

    finally:
        Path(xlsx_path).unlink()


def test_streaming_xlsx_reader_with_shared_strings() -> None:
    """Test stream_rows with shared strings."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        # Add text values that will use shared strings
        ws.append(["Header1", "Header2", "Header3"])
        ws.append(["Value1", "Value2", "Value3"])
        ws.append(["Value1", "Value2", "Value3"])  # Duplicate values
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)

        # First extract shared strings
        from xlsx_streamer.xlsx_metadata_extractor import XLSXMetadataExtractor

        extractor = XLSXMetadataExtractor(source)
        shared_strings, worksheet_path = extractor.extract_metadata(None)

        # Now stream rows with shared strings
        reader = StreamingXlsxReader(source, target_worksheet_filepath=worksheet_path)
        rows = list(reader.stream_rows(shared_strings))

        assert len(rows) >= 1  # At least one row

    finally:
        Path(xlsx_path).unlink()


def test_streaming_xlsx_reader_with_numbers() -> None:
    """Test stream_rows correctly handles numeric values."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([1, 2.5, 100])
        ws.append([42, 3.14159, -5])
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        reader = StreamingXlsxReader(source, target_worksheet_filepath="xl/worksheets/sheet1.xml")

        rows = list(reader.stream_rows([]))
        assert len(rows) == 2

        # Check numeric values
        assert rows[0][0] == 1
        assert isinstance(rows[0][1], float)
        assert rows[0][2] == 100

    finally:
        Path(xlsx_path).unlink()


def test_streaming_xlsx_reader_with_empty_cells() -> None:
    """Test stream_rows handles empty cells correctly."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "A"
        ws["C1"] = "C"  # Skip B1
        ws["A2"] = "X"
        ws["B2"] = "Y"
        ws["D2"] = "Z"  # Skip C2
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        reader = StreamingXlsxReader(source, target_worksheet_filepath="xl/worksheets/sheet1.xml")

        rows = list(reader.stream_rows([]))
        assert len(rows) == 2

        # Check that empty cells are padded with empty strings
        assert rows[0][0] == "A"
        assert rows[0][1] == ""  # B1 is empty
        assert rows[0][2] == "C"

    finally:
        Path(xlsx_path).unlink()


def test_sparse_to_dense_row() -> None:
    """Test _sparse_to_dense_row conversion."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        reader = StreamingXlsxReader(source)

        # Test normal sparse row
        sparse_row = {0: "A", 2: "C", 4: "E"}
        dense_row = reader._sparse_to_dense_row(sparse_row)
        assert dense_row == ["A", "", "C", "", "E"]

        # Test empty sparse row
        dense_row = reader._sparse_to_dense_row({})
        assert dense_row == []

        # Test single cell
        dense_row = reader._sparse_to_dense_row({3: "D"})
        assert dense_row == ["", "", "", "D"]

    finally:
        Path(temp_path).unlink()


def test_address_to_index() -> None:
    """Test _address_to_index conversion."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        reader = StreamingXlsxReader(source)

        # Test simple addresses
        assert reader._address_to_index("A1") == 0
        assert reader._address_to_index("B1") == 1
        assert reader._address_to_index("C1") == 2
        assert reader._address_to_index("Z1") == 25

        # Test multi-letter addresses
        assert reader._address_to_index("AA1") == 26
        assert reader._address_to_index("AB1") == 27
        assert reader._address_to_index("AZ1") == 51
        assert reader._address_to_index("BA1") == 52

        # Test different row numbers
        assert reader._address_to_index("A100") == 0
        assert reader._address_to_index("B99") == 1

        # Test lowercase
        assert reader._address_to_index("a1") == 0
        assert reader._address_to_index("z1") == 25

    finally:
        Path(temp_path).unlink()


def test_iterable_to_file_read() -> None:
    """Test IterableToFile.read() method."""

    def chunk_iter() -> Iterator[bytes]:
        yield b"Hello"
        yield b"World"
        yield b"Test"

    file_obj = IterableToFile(chunk_iter())

    # Read 5 bytes
    data = file_obj.read(5)
    assert data == b"Hello"

    # Read 5 more bytes
    data = file_obj.read(5)
    assert data == b"World"

    # Read remaining
    data = file_obj.read(-1)
    assert data == b"Test"


def test_iterable_to_file_read_all() -> None:
    """Test IterableToFile.read() with no size argument."""

    def chunk_iter() -> Iterator[bytes]:
        yield b"Hello"
        yield b"World"

    file_obj = IterableToFile(chunk_iter())

    # Read all at once
    data = file_obj.read()
    assert data == b"HelloWorld"


def test_iterable_to_file_read_larger_than_available() -> None:
    """Test IterableToFile.read() when requesting more than available."""

    def chunk_iter() -> Iterator[bytes]:
        yield b"Hello"

    file_obj = IterableToFile(chunk_iter())

    # Request more than available
    data = file_obj.read(100)
    assert data == b"Hello"


def test_iterable_to_file_read_zero() -> None:
    """Test IterableToFile.read(0) returns empty bytes."""

    def chunk_iter() -> Iterator[bytes]:
        yield b"Hello"

    file_obj = IterableToFile(chunk_iter())

    # Read 0 bytes
    data = file_obj.read(0)
    assert data == b""

    # Next read should still work
    data = file_obj.read(5)
    assert data == b"Hello"


def test_iterable_to_file_empty_iterator() -> None:
    """Test IterableToFile with empty iterator."""

    def chunk_iter() -> Iterator[bytes]:
        return
        yield  # pragma: no cover

    file_obj = IterableToFile(chunk_iter())

    # Read from empty iterator
    data = file_obj.read()
    assert data == b""


def test_streaming_xlsx_reader_enum() -> None:
    """Test StreamingXlsxReader enum values."""
    assert StreamingXlsxReader.XlsxFilePaths.DEFAULT_WORKSHEET.value == "xl/worksheets/sheet1.xml"


def test_streaming_xlsx_reader_with_inline_strings() -> None:
    """Test stream_rows with inline strings (inlineStr type)."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        # Create XLSX file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Test1", "Test2"])
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        reader = StreamingXlsxReader(source, target_worksheet_filepath="xl/worksheets/sheet1.xml")

        # Stream rows
        rows = list(reader.stream_rows([]))
        assert len(rows) >= 1

    finally:
        Path(xlsx_path).unlink()


def test_streaming_xlsx_reader_stream_error_handling() -> None:
    """Test stream_rows error handling with invalid XLSX."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"not a valid xlsx file")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        reader = StreamingXlsxReader(source, target_worksheet_filepath="xl/worksheets/sheet1.xml")

        # Should raise OSError with any error message
        with pytest.raises(OSError, match="Failed to stream rows"):
            list(reader.stream_rows([]))

    finally:
        # On Windows, file may still be locked
        with contextlib.suppress(PermissionError):
            Path(temp_path).unlink()
