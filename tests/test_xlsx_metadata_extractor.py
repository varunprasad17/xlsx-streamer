"""Tests for XLSXMetadataExtractor."""

from collections.abc import Iterator
from pathlib import Path
import tempfile

import openpyxl
import pytest

from xlsx_streamer.sources.local import LocalFileSource
from xlsx_streamer.xlsx_metadata_extractor import XLSXMetadataExtractor


def test_metadata_extractor_initialization() -> None:
    """Test XLSXMetadataExtractor initialization."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)
        assert extractor.source is source
    finally:
        Path(temp_path).unlink()


def test_metadata_extractor_with_real_xlsx() -> None:
    """Test XLSXMetadataExtractor with real XLSX file."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        # Create a real XLSX file with multiple sheets
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A", "B", "C"])
        ws1.append([1, 2, 3])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["X", "Y", "Z"])
        ws2.append([4, 5, 6])

        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        extractor = XLSXMetadataExtractor(source)

        # Test extracting first sheet (default)
        shared_strings, worksheet_path = extractor.extract_metadata(None)
        assert isinstance(shared_strings, list)
        assert "xl/worksheets/sheet" in worksheet_path

        # Test extracting specific sheet by name
        shared_strings2, worksheet_path2 = extractor.extract_metadata("Sheet2")
        assert isinstance(shared_strings2, list)
        assert "xl/worksheets/sheet" in worksheet_path2

    finally:
        Path(xlsx_path).unlink()


def test_metadata_extractor_with_shared_strings() -> None:
    """Test XLSXMetadataExtractor extracts shared strings correctly."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        # Create XLSX with text that will use shared strings
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header1", "Header2", "Header3"])
        ws.append(["Value1", "Value2", "Value3"])
        ws.append(["Value1", "Value2", "Value3"])  # Repeated values use shared strings
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        extractor = XLSXMetadataExtractor(source)

        shared_strings, _ = extractor.extract_metadata(None)
        # Should have extracted shared strings (if Excel uses them)
        assert isinstance(shared_strings, list)

    finally:
        Path(xlsx_path).unlink()


def test_metadata_extractor_invalid_sheet_name() -> None:
    """Test XLSXMetadataExtractor with invalid sheet name raises ValueError."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["A", "B"])
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        extractor = XLSXMetadataExtractor(source)

        # Should raise ValueError for non-existent sheet
        with pytest.raises(ValueError, match="Sheet 'NonExistent' not found"):
            extractor.extract_metadata("NonExistent")

    finally:
        Path(xlsx_path).unlink()


def test_metadata_extractor_default_first_sheet() -> None:
    """Test that None sheet_name uses default first sheet."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        wb.save(xlsx_path)

        source = LocalFileSource(xlsx_path)
        extractor = XLSXMetadataExtractor(source)

        _shared_strings, worksheet_path = extractor.extract_metadata(None)
        assert worksheet_path == "xl/worksheets/sheet1.xml"

    finally:
        Path(xlsx_path).unlink()


def test_parse_workbook_xml_with_valid_data() -> None:
    """Test _parse_workbook_xml with valid XML data."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        # Create valid workbook.xml content
        xml_content = b"""<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
    <sheets>
        <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
        <sheet name="Sheet2" sheetId="2" r:id="rId2"/>
    </sheets>
</workbook>"""

        r_id = extractor._parse_workbook_xml([xml_content], "Sheet2")
        assert r_id == "rId2"

        r_id = extractor._parse_workbook_xml([xml_content], "Sheet1")
        assert r_id == "rId1"

    finally:
        Path(temp_path).unlink()


def test_parse_workbook_xml_with_empty_chunks() -> None:
    """Test _parse_workbook_xml with empty chunks returns None."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        result = extractor._parse_workbook_xml([], "Sheet1")
        assert result is None

    finally:
        Path(temp_path).unlink()


def test_parse_workbook_xml_with_invalid_xml() -> None:
    """Test _parse_workbook_xml with invalid XML returns None."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        result = extractor._parse_workbook_xml([b"<invalid>xml"], "Sheet1")
        assert result is None

    finally:
        Path(temp_path).unlink()


def test_parse_workbook_xml_sheet_not_found() -> None:
    """Test _parse_workbook_xml when sheet name not found returns None."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        xml_content = b"""<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
    <sheets>
        <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
    </sheets>
</workbook>"""

        result = extractor._parse_workbook_xml([xml_content], "NonExistent")
        assert result is None

    finally:
        Path(temp_path).unlink()


def test_parse_rels_xml_with_valid_data() -> None:
    """Test _parse_rels_xml with valid XML data."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        xml_content = b"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
    <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
                  Target="worksheets/sheet1.xml"/>
    <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
                  Target="worksheets/sheet2.xml"/>
</Relationships>"""

        target = extractor._parse_rels_xml([xml_content], "rId2")
        assert target == "worksheets/sheet2.xml"

    finally:
        Path(temp_path).unlink()


def test_parse_rels_xml_with_empty_chunks() -> None:
    """Test _parse_rels_xml with empty chunks returns None."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        result = extractor._parse_rels_xml([], "rId1")
        assert result is None

    finally:
        Path(temp_path).unlink()


def test_parse_rels_xml_with_invalid_xml() -> None:
    """Test _parse_rels_xml with invalid XML returns None."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        result = extractor._parse_rels_xml([b"<invalid>xml"], "rId1")
        assert result is None

    finally:
        Path(temp_path).unlink()


def test_parse_rels_xml_rid_not_found() -> None:
    """Test _parse_rels_xml when r:id not found returns None."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        xml_content = b"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
    <Relationship Id="rId1" Type="type" Target="target.xml"/>
</Relationships>"""

        result = extractor._parse_rels_xml([xml_content], "rId999")
        assert result is None

    finally:
        Path(temp_path).unlink()


def test_parse_shared_strings_xml() -> None:
    """Test _parse_shared_strings_xml with valid data."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        xml_content = b"""<?xml version="1.0" encoding="UTF-8"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="3">
    <si><t>Hello</t></si>
    <si><t>World</t></si>
    <si><t>Test</t></si>
</sst>"""

        def chunk_iter() -> Iterator[bytes]:
            yield xml_content

        shared_strings = extractor._parse_shared_strings_xml(chunk_iter())
        assert shared_strings == ["Hello", "World", "Test"]

    finally:
        Path(temp_path).unlink()


def test_parse_shared_strings_xml_with_rich_text() -> None:
    """Test _parse_shared_strings_xml with rich text (multiple t elements)."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        xml_content = b"""<?xml version="1.0" encoding="UTF-8"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
    <si>
        <r><t>Bold</t></r>
        <r><t>Normal</t></r>
    </si>
</sst>"""

        def chunk_iter() -> Iterator[bytes]:
            yield xml_content

        shared_strings = extractor._parse_shared_strings_xml(chunk_iter())
        assert shared_strings == ["BoldNormal"]

    finally:
        Path(temp_path).unlink()


def test_parse_shared_strings_xml_with_invalid_xml() -> None:
    """Test _parse_shared_strings_xml with invalid XML returns empty list."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        def chunk_iter() -> Iterator[bytes]:
            yield b"<invalid>xml<"

        shared_strings = extractor._parse_shared_strings_xml(chunk_iter())
        assert shared_strings == []

    finally:
        Path(temp_path).unlink()


def test_parse_shared_strings_xml_empty() -> None:
    """Test _parse_shared_strings_xml with empty content."""
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".xlsx") as f:
        f.write(b"test")
        temp_path = f.name

    try:
        source = LocalFileSource(temp_path)
        extractor = XLSXMetadataExtractor(source)

        def chunk_iter() -> Iterator[bytes]:
            yield b""

        shared_strings = extractor._parse_shared_strings_xml(chunk_iter())
        assert shared_strings == []

    finally:
        Path(temp_path).unlink()


def test_metadata_extractor_enums() -> None:
    """Test that enum values are correct."""
    assert XLSXMetadataExtractor.XlsxMetadataPaths.WORKBOOK.value == "xl/workbook.xml"
    assert XLSXMetadataExtractor.XlsxMetadataPaths.SHARED_STRINGS.value == "xl/sharedStrings.xml"
    assert (
        XLSXMetadataExtractor.XlsxMetadataPaths.WORKBOOK_RELS.value == "xl/_rels/workbook.xml.rels"
    )
    assert (
        XLSXMetadataExtractor.XlsxMetadataPaths.DEFAULT_WORKSHEET.value
        == "xl/worksheets/sheet1.xml"
    )

    assert (
        XLSXMetadataExtractor.XlsxNamespaces.MAIN.value
        == "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    )
    assert (
        XLSXMetadataExtractor.XlsxNamespaces.REL.value
        == "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    assert (
        XLSXMetadataExtractor.XlsxNamespaces.PACKAGE_REL.value
        == "http://schemas.openxmlformats.org/package/2006/relationships"
    )
