import contextlib
from pathlib import Path
import tempfile

import openpyxl
from typer.testing import CliRunner

from xlsx_streamer.cli import app


def test_cli_stream_xlsx_to_csv() -> None:
    """Test CLI streaming XLSX to CSV."""
    # Create a temporary XLSX file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "C"])
    ws.append([1, 2, 3])
    wb.save(xlsx_path)
    try:
        runner = CliRunner()
        result = runner.invoke(app, [xlsx_path])
        assert result.exit_code == 0
        assert "A,B,C" in result.output
        assert "1,2,3" in result.output
    finally:
        Path(xlsx_path).unlink()


def test_cli_with_output_file() -> None:
    """Test CLI with output file option."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as out:
        csv_path = out.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header1", "Header2"])
        ws.append([1, 2])
        wb.save(xlsx_path)

        runner = CliRunner()
        result = runner.invoke(app, [xlsx_path, "--output", csv_path])
        assert result.exit_code == 0
        assert "CSV written to:" in result.output

        # Verify output file was created and has content
        assert Path(csv_path).exists()
        content = Path(csv_path).read_text()
        assert "Header1,Header2" in content

    finally:
        Path(xlsx_path).unlink()
        if Path(csv_path).exists():
            Path(csv_path).unlink()


def test_cli_with_sheet_name() -> None:
    """Test CLI with specific sheet name."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A", "B"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["X", "Y", "Z"])
        ws2.append([1, 2, 3])
        wb.save(xlsx_path)

        runner = CliRunner()
        result = runner.invoke(app, [xlsx_path, "--sheet-name", "Sheet2"])
        assert result.exit_code == 0
        assert "X,Y,Z" in result.output
        assert "1,2,3" in result.output

    finally:
        Path(xlsx_path).unlink()


def test_cli_with_verbose() -> None:
    """Test CLI with verbose logging."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A"])
        wb.save(xlsx_path)

        runner = CliRunner()
        result = runner.invoke(app, [xlsx_path, "--verbose"])
        assert result.exit_code == 0

    finally:
        Path(xlsx_path).unlink()


def test_cli_file_not_found() -> None:
    """Test CLI with non-existent file."""
    runner = CliRunner()
    result = runner.invoke(app, ["/nonexistent/file.xlsx"])
    assert result.exit_code == 1
    assert "Error" in result.output


def test_cli_invalid_sheet_name() -> None:
    """Test CLI with invalid sheet name."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        wb.save(xlsx_path)

        runner = CliRunner()
        result = runner.invoke(app, [xlsx_path, "--sheet-name", "NonExistent"])
        assert result.exit_code == 1
        assert "Error" in result.output

    finally:
        Path(xlsx_path).unlink()


def test_cli_with_invalid_file() -> None:
    """Test CLI with invalid XLSX file."""
    with tempfile.NamedTemporaryFile(mode="wb", suffix=".xlsx", delete=False) as tmp:
        tmp.write(b"not a valid xlsx file")
        invalid_path = tmp.name

    try:
        runner = CliRunner()
        result = runner.invoke(app, [invalid_path])
        assert result.exit_code == 1
        assert "Error" in result.output

    finally:
        # On Windows, file may still be locked
        with contextlib.suppress(PermissionError):
            Path(invalid_path).unlink()


def test_cli_verbose_with_error() -> None:
    """Test CLI verbose mode shows traceback on error."""
    runner = CliRunner()
    result = runner.invoke(app, ["/nonexistent/file.xlsx", "--verbose"])
    assert result.exit_code == 1
    assert "Error" in result.output
